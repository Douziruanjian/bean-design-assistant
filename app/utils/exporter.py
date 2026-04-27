"""
数据导出模块

支持将工单、报价单、客户数据导出为 Excel 或 CSV 格式
"""
import os
from typing import List, Optional
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from ..database.models import Order, Quotation, Customer
from ..database.db_manager import DatabaseManager


class DataExporter:
    """数据导出器"""
    
    def __init__(self, db: DatabaseManager):
        """
        初始化导出器
        
        Args:
            db: 数据库管理器实例
        """
        self.db = db
    
    def _create_header_style(self):
        """创建表头样式"""
        if not HAS_OPENPYXL:
            return None
        return Font(bold=True, size=12)
    
    def _create_cell_style(self):
        """创建单元格样式"""
        if not HAS_OPENPYXL:
            return None
        return Alignment(wrap_text=True)
    
    def export_orders_to_excel(self, output_path: str, 
                                status: Optional[str] = None,
                                start_date: Optional[str] = None,
                                end_date: Optional[str] = None) -> bool:
        """
        导出工单到 Excel
        
        Args:
            output_path: 输出文件路径
            status: 工单状态筛选
            start_date: 开始日期
            end_date: 结束日期
            
        Returns:
            bool: 是否成功
        """
        if not HAS_OPENPYXL:
            return False
        
        orders = self.db.get_orders(status=status, start_date=start_date, end_date=end_date)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "工单列表"
        
        # 表头
        headers = ['工单号', '客户名称', '客户电话', '工单描述', '金额', '状态', '创建时间', '更新时间']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self._create_header_style()
            cell.alignment = Alignment(horizontal="center")
        
        # 数据行
        status_map = {
            'pending': '待处理',
            'in_progress': '进行中',
            'completed': '已完成',
            'cancelled': '已取消'
        }
        
        for row_idx, order in enumerate(orders, 2):
            ws.cell(row=row_idx, column=1, value=order.order_no)
            ws.cell(row=row_idx, column=2, value=order.customer_name)
            ws.cell(row=row_idx, column=3, value=order.customer_phone)
            ws.cell(row=row_idx, column=4, value=order.description)
            ws.cell(row=row_idx, column=5, value=order.total_amount)
            ws.cell(row=row_idx, column=6, value=status_map.get(order.status, order.status))
            ws.cell(row=row_idx, column=7, value=order.created_at)
            ws.cell(row=row_idx, column=8, value=order.updated_at)
        
        # 调整列宽
        column_widths = [18, 15, 15, 30, 12, 12, 20, 20]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width
        
        # 保存文件
        wb.save(output_path)
        return True
    
    def export_quotations_to_excel(self, output_path: str,
                                    status: Optional[str] = None) -> bool:
        """
        导出报价单到 Excel
        
        Args:
            output_path: 输出文件路径
            status: 报价单状态筛选
            
        Returns:
            bool: 是否成功
        """
        if not HAS_OPENPYXL:
            return False
        
        quotations = self.db.get_quotations(status=status)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "报价单列表"
        
        # 表头
        headers = ['报价单号', '客户名称', '项目数量', '总金额', '有效期至', '状态', '创建时间']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self._create_header_style()
            cell.alignment = Alignment(horizontal="center")
        
        # 数据行
        status_map = {
            'draft': '草稿',
            'confirmed': '已确认',
            'expired': '已过期'
        }
        
        for row_idx, quotation in enumerate(quotations, 2):
            ws.cell(row=row_idx, column=1, value=quotation.quotation_no)
            ws.cell(row=row_idx, column=2, value=quotation.customer_name)
            ws.cell(row=row_idx, column=3, value=len(quotation.items))
            ws.cell(row=row_idx, column=4, value=quotation.total_amount)
            ws.cell(row=row_idx, column=5, value=quotation.valid_until)
            ws.cell(row=row_idx, column=6, value=status_map.get(quotation.status, quotation.status))
            ws.cell(row=row_idx, column=7, value=quotation.created_at)
        
        # 调整列宽
        column_widths = [18, 15, 12, 12, 15, 12, 20]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width
        
        wb.save(output_path)
        return True
    
    def export_customers_to_excel(self, output_path: str) -> bool:
        """
        导出客户列表到 Excel
        
        Args:
            output_path: 输出文件路径
            
        Returns:
            bool: 是否成功
        """
        if not HAS_OPENPYXL:
            return False
        
        customers = self.db.get_customers()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "客户列表"
        
        # 表头
        headers = ['客户名称', '电话', '地址', '备注', '订单数', '累计消费', '创建时间']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self._create_header_style()
            cell.alignment = Alignment(horizontal="center")
        
        # 数据行
        for row_idx, customer in enumerate(customers, 2):
            ws.cell(row=row_idx, column=1, value=customer.name)
            ws.cell(row=row_idx, column=2, value=customer.phone)
            ws.cell(row=row_idx, column=3, value=customer.address)
            ws.cell(row=row_idx, column=4, value=customer.notes)
            ws.cell(row=row_idx, column=5, value=customer.total_orders)
            ws.cell(row=row_idx, column=6, value=customer.total_spent)
            ws.cell(row=row_idx, column=7, value=customer.created_at)
        
        # 调整列宽
        column_widths = [15, 15, 25, 30, 10, 12, 20]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width
        
        wb.save(output_path)
        return True
    
    def export_orders_to_csv(self, output_path: str,
                              status: Optional[str] = None,
                              start_date: Optional[str] = None,
                              end_date: Optional[str] = None) -> bool:
        """
        导出工单到 CSV
        
        Args:
            output_path: 输出文件路径
            status: 工单状态筛选
            start_date: 开始日期
            end_date: 结束日期
            
        Returns:
            bool: 是否成功
        """
        orders = self.db.get_orders(status=status, start_date=start_date, end_date=end_date)
        
        with open(output_path, 'w', encoding='utf-8') as f:
            # 表头
            f.write("工单号，客户名称，客户电话，工单描述，金额，状态，创建时间，更新时间\n")
            
            # 数据行
            for order in orders:
                f.write(f"{order.order_no},{order.customer_name},{order.customer_phone},"
                       f"{order.description},{order.total_amount},{order.status},"
                       f"{order.created_at},{order.updated_at}\n")
        
        return True
    
    def export_quotation_detail_to_excel(self, quotation_id: int, 
                                          output_path: str) -> bool:
        """
        导出单个报价单详情到 Excel
        
        Args:
            quotation_id: 报价单 ID
            output_path: 输出文件路径
            
        Returns:
            bool: 是否成功
        """
        if not HAS_OPENPYXL:
            return False
        
        quotation = self.db.get_quotation(quotation_id)
        if not quotation:
            return False
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"报价单-{quotation.quotation_no}"
        
        # 基本信息
        ws.merge_cells('A1:B1')
        ws.cell(row=1, column=1, value=f"报价单号：{quotation.quotation_no}")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        
        ws.cell(row=2, column=1, value=f"客户名称：{quotation.customer_name}")
        ws.cell(row=3, column=1, value=f"有效期至：{quotation.valid_until}")
        ws.cell(row=4, column=1, value=f"状态：{quotation.status}")
        
        # 项目表格表头
        headers = ['项目名称', '数量', '单价', '金额']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        # 项目数据
        for row_idx, item in enumerate(quotation.items, 7):
            ws.cell(row=row_idx, column=1, value=item.name)
            ws.cell(row=row_idx, column=2, value=item.qty)
            ws.cell(row=row_idx, column=3, value=item.unit_price)
            ws.cell(row=row_idx, column=4, value=item.amount)
        
        # 合计
        total_row = 7 + len(quotation.items)
        ws.cell(row=total_row, column=3, value="合计:")
        ws.cell(row=total_row, column=3).font = Font(bold=True)
        ws.cell(row=total_row, column=4, value=quotation.total_amount)
        ws.cell(row=total_row, column=4).font = Font(bold=True)
        
        wb.save(output_path)
        return True
